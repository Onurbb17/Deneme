import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import os
import json
import webbrowser
import time
import threading

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# --- Gradyan arka plan için yardımcı fonksiyon ---
def create_gradient(canvas, width, height, color1, color2):
    limit = height
    (r1, g1, b1) = canvas.winfo_rgb(color1)
    (r2, g2, b2) = canvas.winfo_rgb(color2)
    r_ratio = float(r2 - r1) / limit
    g_ratio = float(g2 - g1) / limit
    b_ratio = float(b2 - b1) / limit

    for i in range(limit):
        nr = int(r1 + (r_ratio * i))
        ng = int(g1 + (g_ratio * i))
        nb = int(b1 + (b_ratio * i))
        color = "#%04x%04x%04x" % (nr, ng, nb)
        canvas.create_line(0, i, width, i, fill=color, tags=("grad",))

# --- Kategori Yönetimi ---
KATEGORI_DOSYASI = "kategoriler.json"

def kategorileri_yukle():
    if os.path.exists(KATEGORI_DOSYASI):
        with open(KATEGORI_DOSYASI, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def kategorileri_kaydet():
    with open(KATEGORI_DOSYASI, "w", encoding="utf-8") as f:
        json.dump(kategoriler, f, ensure_ascii=False, indent=2)

def kategori_degisti(event=None):
    kategori = kategori_var.get()
    listbox.delete(0, tk.END)
    for link in kategoriler.get(kategori, []):
        listbox.insert(tk.END, link)
    etiket["text"] = f"Toplam {len(kategoriler.get(kategori, []))} link var"

def kategori_ekle():
    yeni_kat = simpledialog.askstring("Yeni Kategori", "Kategori Adı:")
    if yeni_kat and yeni_kat not in kategoriler:
        kategoriler[yeni_kat] = []
        kategori_menu["values"] = list(kategoriler.keys())
        kategori_var.set(yeni_kat)
        kategori_degisti()
        kategorileri_kaydet()

def kategori_sil():
    silKat = kategori_var.get()
    if messagebox.askyesno("Kategori Sil", f"{silKat} kategorisini silmek istiyor musun?"):
        kategoriler.pop(silKat, None)
        kategori_menu["values"] = list(kategoriler.keys())
        if kategoriler:
            kategori_var.set(list(kategoriler.keys())[0])
        else:
            kategori_var.set("")
        kategori_degisti()
        kategorileri_kaydet()

def link_ekle():
    yeni_link = entry.get().strip()
    kat = kategori_var.get()
    if yeni_link and kat:
        kategoriler[kat].append(yeni_link)
        listbox.insert(tk.END, yeni_link)
        entry.delete(0, tk.END)
        kategorileri_kaydet()
        kategori_degisti()

def link_sil():
    kat = kategori_var.get()
    secili = listbox.curselection()
    for i in reversed(secili):
        kategoriler[kat].pop(i)
        listbox.delete(i)
    kategorileri_kaydet()
    kategori_degisti()

def dosyadan_link_yukle():
    kat = kategori_var.get()
    dosya_yolu = filedialog.askopenfilename(filetypes=[("Metin Dosyası", "*.txt")])
    if dosya_yolu and kat:
        with open(dosya_yolu, "r", encoding="utf-8") as f:
            for satir in f:
                link = satir.strip()
                if link:
                    kategoriler[kat].append(link)
        kategori_degisti()
        kategorileri_kaydet()

def kaydet():
    kat = kategori_var.get()
    dosya_yolu = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Metin Dosyası", "*.txt")])
    if dosya_yolu and kat:
        with open(dosya_yolu, "w", encoding="utf-8") as f:
            for link in kategoriler[kat]:
                f.write(link + "\n")

def hatali_linkleri_sil():
    kat = kategori_var.get()
    silinecek = []
    for i, link in enumerate(kategoriler[kat]):
        if "HATA" in link:
            silinecek.append(i)
    for i in reversed(silinecek):
        kategoriler[kat].pop(i)
    kategori_degisti()
    kategorileri_kaydet()
    if silinecek:
        messagebox.showinfo("Bilgi", f"{len(silinecek)} hatalı link silindi.")
    else:
        messagebox.showinfo("Bilgi", "Hatalı link bulunamadı.")

# --- Tooltip (Hover Bilgi Balonu) ---
class ToolTip(object):
    def __init__(self, widget, text='widget info'):
        self.waittime = 500     # miliseconds
        self.wraplength = 320   # pixels
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.on_enter)
        self.widget.bind("<Leave>", self.on_leave)
        self.widget.bind("<ButtonPress>", self.on_leave)
        self.id = None
        self.tw = None

    def on_enter(self, event=None):
        self.schedule()

    def on_leave(self, event=None):
        self.unschedule()
        self.hide_tip()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(self.waittime, self.show_tip)

    def unschedule(self):
        _id = self.id
        self.id = None
        if _id:
            self.widget.after_cancel(_id)

    def show_tip(self, event=None):
        x = y = 0
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 30
        y += self.widget.winfo_rooty() + 32
        self.tw = tk.Toplevel(self.widget)
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(self.tw, text=self.text, justify='left',
                       background="#fcf6ff", fg="#333", relief='solid', borderwidth=1,
                       font=("Calibri", "10", "normal"), wraplength = self.wraplength)
        label.pack(ipadx=8, ipady=6)

    def hide_tip(self):
        tw = self.tw
        self.tw= None
        if tw:
            tw.destroy()

# ------------------- FİYAT ÇEKME FONKSİYONU -------------------
def export_to_excel_thematic(df, excel_path):
    df.to_excel(excel_path, index=False)
    wb = load_workbook(excel_path)
    ws = wb.active

    # Tema renkleri (ARGB formatı, başına FF ekle!)
    header_fill = PatternFill(start_color="FF8e2de2", end_color="FF8e2de2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF", name="Calibri", size=12)
    center_align = Alignment(horizontal="center", vertical="center")

    row_fill1 = PatternFill(start_color="FFede7f6", end_color="FFede7f6", fill_type="solid")  # Açık mor
    row_fill2 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    # Başlıklar
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Satırlar için zebra efekti
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = row_fill1 if i % 2 == 0 else row_fill2
        for cell in row:
            cell.fill = fill
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(vertical="center")
    
    # Otomatik sütun genişliği
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # A, B, C...
        for cell in col:
            try:
                val = str(cell.value)
                if val is None: val = ""
                if len(val) > max_length:
                    max_length = len(val)
            except:
                pass
        new_width = min(max_length + 2, 60)  # 60 karakterden uzun olmasın
        ws.column_dimensions[column].width = new_width

    wb.save(excel_path)

def fiyatlari_cek_ve_kaydet():
    kat = kategori_var.get()
    links = kategoriler.get(kat, [])
    if not links:
        messagebox.showwarning("Uyarı", "Çekilecek link yok!")
        return

    urunler, fiyatlar, magazalar, linkler = [], [], [], []
    toplam = len(links)
    progress_var.set(0)
    progress["maximum"] = toplam

    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")
    service = Service("C:\\Users\\ONUR\\Desktop\\Yeni klasör (4)\\chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=chrome_options)

    for idx, url in enumerate(links, start=1):
        driver.get(url)
        time.sleep(5)
        try:
            urun_adi = driver.find_element(By.TAG_NAME, "h1").text.strip()
        except Exception:
            urun_adi = "Ürün adı bulunamadı"
        fiyat = "Fiyat bulunamadı"
        magaza = "Mağaza bulunamadı"
        try:
            # Modern Akakçe'de satıcılar
            satici_listesi = driver.find_element(By.CSS_SELECTOR, "ul#PL")
            ilk_li = satici_listesi.find_elements(By.TAG_NAME, "li")[0]
            a_tag = ilk_li.find_element(By.CSS_SELECTOR, "a.iC.xt_v8")
            fiyat = a_tag.find_element(By.CSS_SELECTOR, ".pt_v8").text.strip().replace("\n", "").replace("\r", "")
            try:
                magaza = a_tag.find_element(By.CSS_SELECTOR, ".v_v8 b").text.strip()
            except Exception:
                magaza_text = a_tag.find_element(By.CSS_SELECTOR, ".v_v8").text.strip()
                magaza = magaza_text.split("/")[-1].strip()
        except Exception as e:
            print("Satıcı bulunamadı:", e)
        urunler.append(urun_adi)
        fiyatlar.append(fiyat)
        magazalar.append(magaza)
        linkler.append(url)
        progress_var.set(idx)
        progress_label.config(text=f"{idx}/{toplam} tamamlandı")
        root.update_idletasks()
    driver.quit()

    now_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_path = f"akakce_fiyatlar_{now_str}.xlsx"
    df = pd.DataFrame({
        "Ürün Adı": urunler,
        "En Uygun Fiyat": fiyatlar,
        "En Uygun Mağaza": magazalar,
        "Ürün Linki": linkler
    })
    export_to_excel_thematic(df, excel_path)
    progress_label.config(text="Tamamlandı")
    messagebox.showinfo("Tamamlandı", f"Fiyatlar kaydedildi:\n{excel_path}")

def fiyatlari_cek_ve_kaydet_thread():
    th = threading.Thread(target=fiyatlari_cek_ve_kaydet)
    th.start()

# --- Ana Arayüz ---
root = tk.Tk()
root.title("Akakce Fiyat Çekici - Modern")
root.geometry("1050x650")
root.resizable(False, False)

canvas = tk.Canvas(root, width=1050, height=650, highlightthickness=0)
canvas.place(x=0, y=0, relwidth=1, relheight=1)
create_gradient(canvas, 1050, 650, "#8e2de2", "#4a00e0")

menu_frame = tk.Frame(root, bg="#2d2257", width=210, height=650)
menu_frame.place(x=0, y=0)

app_title = tk.Label(
    menu_frame,
    text="Akakce Çekici",
    font=("Segoe UI Black", 19, "bold"),
    bg="#2d2257",
    fg="#fff"
)
app_title.place(x=19, y=34)

def hover_on(e): e.widget.config(bg="#6a85f1")
def hover_off(e): e.widget.config(bg="#5032a5")

buton_stil = {"font": ("Segoe UI", 11, "bold"), "fg": "#fff", "bg": "#5032a5", "bd":0, "activeforeground":"#fff", "activebackground":"#6a85f1", "relief":"flat", "width":19, "height":2, "cursor":"hand2"}

btn1 = tk.Button(menu_frame, text="Kategori Ekle", command=kategori_ekle, **buton_stil)
btn1.place(x=12, y=110)
btn2 = tk.Button(menu_frame, text="Kategori Sil", command=kategori_sil, **buton_stil)
btn2.place(x=12, y=160)
btn3 = tk.Button(menu_frame, text="Linkleri Kaydet", command=kaydet, **buton_stil)
btn3.place(x=12, y=210)
btn4 = tk.Button(menu_frame, text="Dosyadan Yükle", command=dosyadan_link_yukle, **buton_stil)
btn4.place(x=12, y=260)
btn5 = tk.Button(menu_frame, text="Hatalı Linkleri Sil", command=hatali_linkleri_sil, **buton_stil)
btn5.place(x=12, y=310)

for btn in [btn1, btn2, btn3, btn4, btn5]:
    btn.bind("<Enter>", hover_on)
    btn.bind("<Leave>", hover_off)

txt_tooltip = (
    "Önce bilgisayarınızda bir .txt dosyası oluşturun.\n"
    "Her satıra bir link gelecek şekilde linklerinizi ekleyin.\n"
    "Sonra bu butona tıklayarak o dosyayı seçin ve tüm linkler otomatik olarak eklensin."
)
ToolTip(btn4, txt_tooltip)

def open_github(event=None):
    webbrowser.open_new("https://github.com/Onurbb17")
faturos_label = tk.Label(menu_frame, text="Faturos", font=("Segoe UI", 10, "bold", "underline"), bg="#2d2257", fg="#a6d3ff", cursor="hand2")
faturos_label.place(x=22, y=600)
faturos_label.bind("<Button-1>", open_github)

main_frame = tk.Frame(root, bg="#ffffff", width=820, height=610, highlightbackground="#eee", highlightthickness=0)
main_frame.place(x=210, y=20)

kategoriler = kategorileri_yukle()
kategori_var = tk.StringVar()
if kategoriler:
    kategori_var.set(list(kategoriler.keys())[0])
else:
    kategori_var.set("Genel")
    kategoriler["Genel"] = []

top_entry_frame = tk.Frame(main_frame, bg="#fff")
top_entry_frame.place(x=30, y=25, width=760, height=44)

kategori_label = tk.Label(top_entry_frame, text="Kategori:", font=("Montserrat", 12, "bold"), bg="#fff", fg="#4a00e0")
kategori_label.grid(row=0, column=0, padx=(0,2), pady=0, sticky="w")

kategori_menu = ttk.Combobox(top_entry_frame, textvariable=kategori_var, values=list(kategoriler.keys()), width=20, font=("Montserrat", 12))
kategori_menu.grid(row=0, column=1, padx=(0,12), pady=0, sticky="w")
kategori_menu.bind("<<ComboboxSelected>>", kategori_degisti)

entry = ttk.Entry(top_entry_frame, width=42, font=("Montserrat", 12))
entry.grid(row=0, column=2, padx=(0,8), pady=0, sticky="w")

ekle_btn = tk.Button(top_entry_frame, text="Link Ekle", font=("Montserrat", 11, "bold"), fg="#fff", bg="#8e2de2", bd=0, activebackground="#5032a5", activeforeground="#fff", relief="flat", width=10, height=1, cursor="hand2", command=link_ekle)
ekle_btn.grid(row=0, column=3, padx=(0,0), pady=0)
ekle_btn.bind("<Enter>", lambda e: ekle_btn.config(bg="#5032a5"))
ekle_btn.bind("<Leave>", lambda e: ekle_btn.config(bg="#8e2de2"))

listbox = tk.Listbox(main_frame, width=90, height=22, font=("Calibri", 11), borderwidth=2, relief="groove", selectbackground="#8e2de2", selectforeground="#fff")
listbox.place(x=30, y=80)
scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=listbox.yview)
scrollbar.place(x=785, y=80, height=348)
listbox.config(yscrollcommand=scrollbar.set)

sil_btn = tk.Button(main_frame, text="Seçili Linki Sil", font=("Montserrat", 10, "bold"), fg="#fff", bg="#c471f5", bd=0, activebackground="#a251bf", activeforeground="#fff", relief="flat", width=18, cursor="hand2", command=link_sil)
sil_btn.place(x=30, y=440)
sil_btn.bind("<Enter>", lambda e: sil_btn.config(bg="#a251bf"))
sil_btn.bind("<Leave>", lambda e: sil_btn.config(bg="#c471f5"))

etiket = tk.Label(main_frame, text="Toplam 0 link var", font=("Montserrat", 11, "bold"), bg="#fff", fg="#6a85b6")
etiket.place(x=32, y=490)

magaza_label = tk.Label(main_frame, text="Mağaza Adı (Vurgu):", font=("Montserrat", 11), bg="#fff", fg="#4a00e0")
magaza_label.place(x=210, y=490)
magaza_var = tk.StringVar(value="KendiMağazanız")
magaza_entry = ttk.Entry(main_frame, textvariable=magaza_var, width=15, font=("Verdana", 11))
magaza_entry.place(x=350, y=490)

cek_btn = tk.Button(main_frame, text="Fiyatları Çek ve Excel'e Kaydet", font=("Montserrat", 12, "bold"), fg="#fff", bg="#6a85f1", bd=0, activebackground="#8e2de2", activeforeground="#fff", relief="flat", width=27, height=2, cursor="hand2", command=fiyatlari_cek_ve_kaydet_thread)
cek_btn.place(x=510, y=480)
cek_btn.bind("<Enter>", lambda e: cek_btn.config(bg="#8e2de2"))
cek_btn.bind("<Leave>", lambda e: cek_btn.config(bg="#6a85f1"))

progress_var = tk.DoubleVar()
progress = ttk.Progressbar(main_frame, variable=progress_var, orient="horizontal", length=600, mode="determinate")
progress.place(x=100, y=550)
progress_label = tk.Label(main_frame, text="0/0 tamamlandı", bg="#fff", fg="#4a00e0", font=("Verdana", 10, "italic"))
progress_label.place(x=720, y=547)

kategori_degisti()

root.mainloop()